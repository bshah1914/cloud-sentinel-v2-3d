"""
CloudSentinel Enterprise — Multi-Cloud Security Platform
FastAPI backend with provider plugin architecture for AWS, Azure, and GCP.
"""

import json
import shutil
import threading
import uuid
import yaml
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
from pydantic import BaseModel
from jose import JWTError, jwt
from passlib.context import CryptContext

# ── Paths ────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
ACCOUNT_DATA_DIR = BASE_DIR / "account-data"
CONFIG_FILE = BASE_DIR / "config.json"
CONFIG_DEMO = BASE_DIR / "config.json.demo"
AUDIT_CONFIG = BASE_DIR / "audit_config.yaml"
AUDIT_OVERRIDE = BASE_DIR / "config" / "audit_config_override.yaml"
USERS_FILE = BASE_DIR / "backend" / "users.json"

# ── JWT / Auth Config ────────────────────────────────────────────
SECRET_KEY = "cloudlunar-enterprise-secret-change-in-production-2024"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 hours

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")


# ── User Management ─────────────────────────────────────────────
def _load_users() -> list[dict]:
    if USERS_FILE.exists():
        with open(USERS_FILE) as f:
            return json.load(f)
    return []


def _save_users(users: list[dict]):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=2)


def _init_admin():
    users = _load_users()
    admin = next((u for u in users if u["username"] == "admin"), None)
    if admin:
        admin["hashed_password"] = pwd_context.hash("admin123")
    else:
        users.append({
            "username": "admin",
            "hashed_password": pwd_context.hash("admin123"),
            "role": "admin",
            "created": datetime.now().isoformat(),
        })
    _save_users(users)


def _authenticate_user(username: str, password: str) -> Optional[dict]:
    for user in _load_users():
        if user["username"] == username and pwd_context.verify(password, user["hashed_password"]):
            return user
    return None


def _create_token(data: dict, expires_delta: timedelta) -> str:
    to_encode = data.copy()
    to_encode["exp"] = datetime.utcnow() + expires_delta
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(401, "Invalid token")
    except JWTError:
        raise HTTPException(401, "Invalid or expired token")
    for user in _load_users():
        if user["username"] == username:
            return user
    raise HTTPException(401, "User not found")


# ── App Setup ────────────────────────────────────────────────────
app = FastAPI(title="CloudSentinel Enterprise API", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_init_admin()

# ── Register Provider Plugins ────────────────────────────────────
from providers.registry import registry
from providers.aws.collector import AWSCollector
from providers.aws.parser import AWSParser
from providers.aws.auditor import AWSAuditor
from providers.aws.routes import AWSRoutes

from providers.azure.collector import AzureCollector
from providers.azure.parser import AzureParser
from providers.azure.auditor import AzureAuditor
from providers.azure.routes import AzureRoutes

from providers.gcp.collector import GCPCollector
from providers.gcp.parser import GCPParser
from providers.gcp.auditor import GCPAuditor
from providers.gcp.routes import GCPRoutes

# Register all providers
registry.register("aws", AWSCollector(), AWSParser(), AWSAuditor(), AWSRoutes())
registry.register("azure", AzureCollector(), AzureParser(), AzureAuditor(), AzureRoutes())
registry.register("gcp", GCPCollector(), GCPParser(), GCPAuditor(), GCPRoutes())

# Mount provider-specific routes
for pid in registry.provider_ids:
    plugin = registry.get(pid)
    plugin.routes.register_routes(app, get_current_user, ACCOUNT_DATA_DIR)


# ── In-memory State ─────────────────────────────────────────────
scan_jobs: dict = {}


# ── Pydantic Models ─────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


class AccountIn(BaseModel):
    id: str
    name: str
    provider: str = "aws"      # aws, azure, gcp
    default: bool = False


class CIDRIn(BaseModel):
    cidr: str
    name: str


class ScanRequestModel(BaseModel):
    account_name: Optional[str] = None
    provider: str = "aws"
    credentials: Optional[dict] = None
    region: Optional[str] = "all"
    # Legacy AWS fields (backward compat)
    aws_access_key_id: Optional[str] = None
    aws_secret_access_key: Optional[str] = None
    aws_region: Optional[str] = None


class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str = "viewer"       # admin, editor, viewer


# ── Helpers ──────────────────────────────────────────────────────
def _load_config() -> dict:
    cfg_path = CONFIG_FILE if CONFIG_FILE.exists() else CONFIG_DEMO
    with open(cfg_path) as f:
        return json.load(f)


def _save_config(config: dict):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path) as f:
        return json.load(f)


def _get_account_dirs(provider: str = None) -> list[dict]:
    """List accounts that have collected data, optionally filtered by provider."""
    results = []
    if not ACCOUNT_DATA_DIR.exists():
        return results

    if provider:
        pdir = ACCOUNT_DATA_DIR / provider
        if pdir.exists():
            for d in pdir.iterdir():
                if d.is_dir() and not d.name.startswith("."):
                    results.append({"name": d.name, "provider": provider})
    else:
        for pdir in ACCOUNT_DATA_DIR.iterdir():
            if pdir.is_dir() and pdir.name in ("aws", "azure", "gcp"):
                for d in pdir.iterdir():
                    if d.is_dir() and not d.name.startswith("."):
                        results.append({"name": d.name, "provider": pdir.name})
        # Legacy: check for accounts not under a provider folder (old AWS format)
        for d in ACCOUNT_DATA_DIR.iterdir():
            if d.is_dir() and d.name not in ("aws", "azure", "gcp") and not d.name.startswith("."):
                results.append({"name": d.name, "provider": "aws"})
    return results


def _get_regions_for_account(account_name: str, provider: str = "aws") -> list[str]:
    acct_dir = ACCOUNT_DATA_DIR / provider / account_name
    # Fallback to legacy path
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        return []
    return [d.name for d in acct_dir.iterdir() if d.is_dir() and not d.name.startswith(".")]


# ── AUTH ENDPOINTS ───────────────────────────────────────────────
@app.post("/api/auth/login")
def login(req: LoginRequest):
    user = _authenticate_user(req.username, req.password)
    if not user:
        raise HTTPException(401, "Invalid username or password")
    token = _create_token(
        {"sub": user["username"], "role": user["role"]},
        timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    return {
        "access_token": token, "token_type": "bearer",
        "user": {"username": user["username"], "role": user["role"]},
    }


@app.get("/api/auth/me")
def get_me(user: dict = Depends(get_current_user)):
    return {"username": user["username"], "role": user["role"], "created": user.get("created")}


# ── USER MANAGEMENT (RBAC) ──────────────────────────────────────
@app.get("/api/users")
def list_users(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return {"users": [
        {"username": u["username"], "role": u["role"], "created": u.get("created")}
        for u in _load_users()
    ]}


@app.post("/api/users")
def create_user(req: UserCreateRequest, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    users = _load_users()
    if any(u["username"] == req.username for u in users):
        raise HTTPException(400, "Username already exists")
    users.append({
        "username": req.username,
        "hashed_password": pwd_context.hash(req.password),
        "role": req.role,
        "created": datetime.now().isoformat(),
    })
    _save_users(users)
    return {"status": "ok", "user": {"username": req.username, "role": req.role}}


@app.delete("/api/users/{username}")
def delete_user(username: str, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    if username == user["username"]:
        raise HTTPException(400, "Cannot delete yourself")
    users = _load_users()
    users = [u for u in users if u["username"] != username]
    _save_users(users)
    return {"status": "ok"}


# ── PROVIDERS ENDPOINT ───────────────────────────────────────────
@app.get("/api/providers")
def list_providers(user: dict = Depends(get_current_user)):
    """List all registered cloud providers with their metadata."""
    return {"providers": registry.list_providers()}


@app.post("/api/providers/{provider_id}/validate")
def validate_provider_credentials(provider_id: str, credentials: dict, user: dict = Depends(get_current_user)):
    """Validate credentials for a specific provider."""
    plugin = registry.get(provider_id)
    if not plugin:
        raise HTTPException(404, f"Provider '{provider_id}' not found")
    return plugin.collector.validate_credentials(credentials)


@app.get("/api/providers/{provider_id}/regions")
def get_provider_regions(provider_id: str, user: dict = Depends(get_current_user)):
    """Get available regions for a provider (static list, no credentials needed)."""
    plugin = registry.get(provider_id)
    if not plugin:
        raise HTTPException(404, f"Provider '{provider_id}' not found")
    return {"regions": plugin.collector.get_regions({})}


# ── ACCOUNTS (Multi-Cloud) ──────────────────────────────────────
@app.get("/api/accounts")
def list_accounts(user: dict = Depends(get_current_user)):
    config = _load_config()
    accounts = config.get("accounts", [])
    enriched = []
    seen = set()

    for acct in accounts:
        provider = acct.get("provider", "aws")
        name = acct["name"]
        has_data = (ACCOUNT_DATA_DIR / provider / name).exists()
        # Legacy check
        if not has_data and provider == "aws":
            has_data = (ACCOUNT_DATA_DIR / name).exists()
        regions = _get_regions_for_account(name, provider) if has_data else []
        enriched.append({**acct, "provider": provider, "has_data": has_data, "regions": regions})
        seen.add(f"{provider}:{name}")

    # Auto-detect accounts with data not in config
    for item in _get_account_dirs():
        key = f"{item['provider']}:{item['name']}"
        if key not in seen:
            regions = _get_regions_for_account(item["name"], item["provider"])
            enriched.append({
                "id": "auto-detected", "name": item["name"],
                "provider": item["provider"],
                "default": len(enriched) == 0, "has_data": True, "regions": regions,
            })

    return {"accounts": enriched}


@app.post("/api/accounts")
def add_account(account: AccountIn, user: dict = Depends(get_current_user)):
    config = _load_config()
    for existing in config.get("accounts", []):
        if existing["id"] == account.id and existing.get("provider", "aws") == account.provider:
            raise HTTPException(400, "Account with this ID already exists for this provider")
    config.setdefault("accounts", []).append(account.dict())
    _save_config(config)
    return {"status": "ok", "account": account.dict()}


@app.delete("/api/accounts/{account_id}")
def remove_account(account_id: str, provider: str = "aws", user: dict = Depends(get_current_user)):
    config = _load_config()
    removed_name = None
    removed_provider = provider
    for acct in config.get("accounts", []):
        if acct["id"] == account_id:
            removed_name = acct["name"]
            removed_provider = acct.get("provider", "aws")
            break

    config["accounts"] = [a for a in config.get("accounts", []) if a["id"] != account_id]
    _save_config(config)

    data_deleted = False
    if removed_name:
        # New path: account-data/{provider}/{name}
        acct_data_dir = ACCOUNT_DATA_DIR / removed_provider / removed_name
        if acct_data_dir.exists():
            shutil.rmtree(acct_data_dir)
            data_deleted = True
        # Legacy path: account-data/{name}
        legacy_dir = ACCOUNT_DATA_DIR / removed_name
        if legacy_dir.exists():
            shutil.rmtree(legacy_dir)
            data_deleted = True

    return {"status": "ok", "account_removed": removed_name, "provider": removed_provider, "data_deleted": data_deleted}


# ── CIDRs ────────────────────────────────────────────────────────
@app.get("/api/cidrs")
def list_cidrs(user: dict = Depends(get_current_user)):
    config = _load_config()
    cidrs = config.get("cidrs", {})
    return {"cidrs": [{"cidr": k, **v} for k, v in cidrs.items()]}


@app.post("/api/cidrs")
def add_cidr(cidr_in: CIDRIn, user: dict = Depends(get_current_user)):
    config = _load_config()
    config.setdefault("cidrs", {})[cidr_in.cidr] = {"name": cidr_in.name}
    _save_config(config)
    return {"status": "ok"}


@app.delete("/api/cidrs/{cidr}")
def remove_cidr(cidr: str, user: dict = Depends(get_current_user)):
    config = _load_config()
    config.get("cidrs", {}).pop(cidr, None)
    _save_config(config)
    return {"status": "ok"}


# ── UNIFIED SCAN (Multi-Cloud) ──────────────────────────────────
@app.post("/api/scan/start")
def start_scan(req: ScanRequestModel, user: dict = Depends(get_current_user)):
    """Start a background scan for any cloud provider."""
    provider_id = req.provider
    plugin = registry.get(provider_id)
    if not plugin:
        raise HTTPException(400, f"Unknown provider: {provider_id}")

    # Build credentials dict
    credentials = req.credentials or {}
    # Legacy AWS fields
    if provider_id == "aws" and not credentials:
        if req.aws_access_key_id:
            credentials = {
                "access_key_id": req.aws_access_key_id,
                "secret_access_key": req.aws_secret_access_key,
            }

    job_id = str(uuid.uuid4())[:8]
    account_name = req.account_name or "default"
    requested_region = req.region or req.aws_region or "all"

    scan_jobs[job_id] = {
        "id": job_id, "provider": provider_id,
        "status": "running", "account": account_name,
        "started": datetime.now().isoformat(),
        "log": [], "progress": 0,
        "regions_scanned": [], "regions_total": 0,
    }

    def _run():
        job = scan_jobs[job_id]
        try:
            collector = plugin.collector

            # Determine regions
            if requested_region == "all":
                job["log"].append(f"Discovering enabled {plugin.info.short_name} regions...")
                regions_to_scan = collector.get_regions(credentials)
                job["log"].append(f"Found {len(regions_to_scan)} regions")
            else:
                regions_to_scan = [requested_region]

            job["regions_total"] = len(regions_to_scan)

            total_collected = 0
            total_errors = []

            for idx, region in enumerate(regions_to_scan):
                job["progress"] = int(5 + (85 * idx / max(len(regions_to_scan), 1)))
                job["log"].append(f"[{idx+1}/{len(regions_to_scan)}] Scanning {region}...")

                result = collector.collect(
                    account_name, region, credentials,
                    ACCOUNT_DATA_DIR, progress_callback=None,
                )

                total_collected += result.get("resources_collected", 0)
                total_errors.extend(result.get("errors", []))
                job["regions_scanned"].append(region)

                if result.get("resources_collected", 0) > 0:
                    job["log"].append(f"  {region}: {result['resources_collected']} resources collected")
                else:
                    job["log"].append(f"  {region}: no resources (or access denied)")

            if total_errors:
                job["log"].append(f"Warnings: {len(total_errors)} total across all regions")

            # Check for any collected data
            acct_dir = ACCOUNT_DATA_DIR / provider_id / account_name
            has_data = acct_dir.exists() and any(acct_dir.rglob("*.json"))

            if not has_data:
                job["status"] = "failed"
                job["log"].append(f"FAILED: No data files created. Check {plugin.info.short_name} credentials.")
                job["completed"] = datetime.now().isoformat()
                return

            job["progress"] = 100
            job["status"] = "completed"
            job["log"].append(f"Scan completed — {total_collected} resources across {len(job['regions_scanned'])} regions")
            job["completed"] = datetime.now().isoformat()

        except Exception as e:
            job["status"] = "failed"
            job["log"].append(f"Error: {str(e)}")
            job["completed"] = datetime.now().isoformat()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    return {"job_id": job_id, "status": "started", "provider": provider_id}


@app.get("/api/scan/status/{job_id}")
def scan_status(job_id: str, user: dict = Depends(get_current_user)):
    if job_id not in scan_jobs:
        raise HTTPException(404, "Job not found")
    return scan_jobs[job_id]


@app.get("/api/scan/history")
def scan_history(user: dict = Depends(get_current_user)):
    return {"jobs": list(scan_jobs.values())}


# ── UNIFIED DASHBOARD ────────────────────────────────────────────
@app.get("/api/dashboard/{provider}/{account_name}")
def dashboard_by_provider(provider: str, account_name: str, user: dict = Depends(get_current_user)):
    """Get dashboard for a specific provider + account."""
    plugin = registry.get(provider)
    if not plugin:
        raise HTTPException(404, f"Provider '{provider}' not found")

    acct_dir = ACCOUNT_DATA_DIR / provider / account_name
    # Legacy fallback for AWS
    if not acct_dir.exists() and provider == "aws":
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        raise HTTPException(404, f"No data for account '{account_name}'")

    return plugin.parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)


# Legacy endpoint (backward compat)
@app.get("/api/dashboard/{account_name}")
def dashboard_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy dashboard endpoint — auto-detects provider."""
    # Check new path first
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            return registry.get(pid).parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
    # Legacy AWS path
    legacy_dir = ACCOUNT_DATA_DIR / account_name
    if legacy_dir.exists():
        return registry.get("aws").parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
    raise HTTPException(404, f"No data for account '{account_name}'")


# ── UNIFIED MULTI-CLOUD OVERVIEW ────────────────────────────────
@app.get("/api/overview")
def multi_cloud_overview(user: dict = Depends(get_current_user)):
    """Aggregated overview across all providers and accounts."""
    overview = {
        "providers": {},
        "total_accounts": 0,
        "total_resources": 0,
        "total_findings": 0,
        "avg_security_score": 0,
    }

    scores = []
    for item in _get_account_dirs():
        pid = item["provider"]
        name = item["name"]
        plugin = registry.get(pid)
        if not plugin:
            continue

        if pid not in overview["providers"]:
            overview["providers"][pid] = {
                "name": plugin.info.name,
                "short_name": plugin.info.short_name,
                "color": plugin.info.color,
                "accounts": [],
                "total_resources": 0,
            }

        try:
            dash = plugin.parser.parse_dashboard(name, ACCOUNT_DATA_DIR)
            total_res = sum(v for v in dash.get("totals", {}).values() if isinstance(v, int))
            findings = plugin.auditor.run_audit(name, ACCOUNT_DATA_DIR)

            acct_info = {
                "name": name, "provider": pid,
                "security_score": dash.get("security_score", 0),
                "total_resources": total_res,
                "total_findings": len(findings),
                "regions_scanned": dash.get("regions_scanned", 0),
            }
            overview["providers"][pid]["accounts"].append(acct_info)
            overview["providers"][pid]["total_resources"] += total_res
            overview["total_accounts"] += 1
            overview["total_resources"] += total_res
            overview["total_findings"] += len(findings)
            scores.append(dash.get("security_score", 0))
        except Exception:
            overview["providers"][pid]["accounts"].append({
                "name": name, "provider": pid, "error": "Failed to load",
            })

    if scores:
        overview["avg_security_score"] = round(sum(scores) / len(scores))

    return overview


# ── LEGACY ENDPOINTS (backward compat for existing frontend) ─────
@app.get("/api/resources/{account_name}")
def get_resources_legacy(account_name: str, user: dict = Depends(get_current_user)):
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            return registry.get(pid).parser.parse_resources(account_name, ACCOUNT_DATA_DIR)
    legacy_dir = ACCOUNT_DATA_DIR / account_name
    if legacy_dir.exists():
        return registry.get("aws").parser.parse_resources(account_name, ACCOUNT_DATA_DIR)
    raise HTTPException(404, f"No data for account '{account_name}'")


@app.get("/api/iam/{account_name}")
def get_iam_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy IAM endpoint — routes to AWS IAM."""
    from providers.aws.parser import AWSParser, _read_json, _get_regions
    acct_dir = ACCOUNT_DATA_DIR / "aws" / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    regions = _get_regions(acct_dir)
    if not regions:
        raise HTTPException(404, "No data found")
    region_dir = acct_dir / regions[0]
    auth = _read_json(region_dir / "iam-get-account-authorization-details.json")
    summary = _read_json(region_dir / "iam-get-account-summary.json")
    users = [{"name": u.get("UserName"), "arn": u.get("Arn"), "created": u.get("CreateDate"),
              "policies": [p["PolicyName"] for p in u.get("AttachedManagedPolicies", [])],
              "inline_policies": list(u.get("UserPolicyList", [])),
              "groups": u.get("GroupList", []), "mfa_devices": u.get("MFADevices", [])}
             for u in auth.get("UserDetailList", [])]
    roles = [{"name": r.get("RoleName"), "arn": r.get("Arn"), "created": r.get("CreateDate"),
              "policies": [p["PolicyName"] for p in r.get("AttachedManagedPolicies", [])],
              "trust_policy": r.get("AssumeRolePolicyDocument")}
             for r in auth.get("RoleDetailList", [])]
    policies = [{"name": p.get("PolicyName"), "arn": p.get("Arn"),
                 "attachment_count": p.get("AttachmentCount"), "is_attachable": p.get("IsAttachable")}
                for p in auth.get("Policies", [])]
    return {"account": account_name, "summary": summary.get("SummaryMap", {}),
            "users": users, "roles": roles, "policies": policies}


@app.get("/api/security-groups/{account_name}")
def get_security_groups_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy SG endpoint — routes to AWS."""
    from providers.aws.parser import _read_json, _get_regions
    acct_dir = ACCOUNT_DATA_DIR / "aws" / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    regions = _get_regions(acct_dir)
    results = []
    for region in regions:
        sgs = _read_json(acct_dir / region / "ec2-describe-security-groups.json")
        for sg in sgs.get("SecurityGroups", []):
            open_rules = []
            for rule in sg.get("IpPermissions", []):
                for ip_range in rule.get("IpRanges", []):
                    if ip_range.get("CidrIp") == "0.0.0.0/0":
                        open_rules.append({"protocol": rule.get("IpProtocol"), "from_port": rule.get("FromPort"), "to_port": rule.get("ToPort")})
                for ip_range in rule.get("Ipv6Ranges", []):
                    if ip_range.get("CidrIpv6") == "::/0":
                        open_rules.append({"protocol": rule.get("IpProtocol"), "from_port": rule.get("FromPort"), "to_port": rule.get("ToPort")})
            if open_rules:
                results.append({"region": region, "group_id": sg["GroupId"], "group_name": sg["GroupName"],
                                "vpc_id": sg.get("VpcId"), "open_rules": open_rules,
                                "severity": "CRITICAL" if any(r.get("from_port") in (22, 3389, 3306, 5432) or r.get("protocol") == "-1" for r in open_rules) else "HIGH"})
    return {"account": account_name, "risky_groups": results}


@app.get("/api/audit/config")
def get_audit_config(user: dict = Depends(get_current_user)):
    try:
        with open(AUDIT_CONFIG) as f:
            cfg = yaml.safe_load(f) or {}
        if AUDIT_OVERRIDE.exists():
            with open(AUDIT_OVERRIDE) as f:
                override = yaml.safe_load(f) or {}
            for k, v in override.items():
                if k not in cfg:
                    cfg[k] = {"title": "Unknown", "severity": "High", "group": "unknown"}
                cfg[k].update(v)
        return {"findings": [{"id": fid, **conf} for fid, conf in cfg.items()]}
    except Exception:
        return {"findings": []}


@app.post("/api/audit/run/{account_name}")
def run_audit_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy audit — auto-detects provider."""
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            plugin = registry.get(pid)
            findings = plugin.auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
            severity_counts = {}
            for f in findings:
                s = f.get("severity", "INFO")
                severity_counts[s] = severity_counts.get(s, 0) + 1
            return {"account": account_name, "provider": pid,
                    "total": len(findings), "findings": findings,
                    "summary": {s: severity_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]}}
    # Legacy AWS
    legacy_dir = ACCOUNT_DATA_DIR / account_name
    if legacy_dir.exists():
        findings = registry.get("aws").auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
        severity_counts = {}
        for f in findings:
            s = f.get("severity", "INFO")
            severity_counts[s] = severity_counts.get(s, 0) + 1
        return {"account": account_name, "provider": "aws",
                "total": len(findings), "findings": findings,
                "summary": {s: severity_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]}}
    raise HTTPException(404, f"No data for account '{account_name}'")


# ── REPORT EXPORT ────────────────────────────────────────────────
from fastapi.responses import Response
from report_generator import (
    generate_dashboard_pdf, generate_dashboard_csv,
    generate_audit_pdf, generate_audit_csv,
)


@app.get("/api/export/dashboard/{provider}/{account_name}")
def export_dashboard(provider: str, account_name: str, format: str = "pdf",
                     user: dict = Depends(get_current_user)):
    """Export dashboard report as PDF or CSV."""
    plugin = registry.get(provider)
    if not plugin:
        raise HTTPException(400, f"Unknown provider '{provider}'")
    data = plugin.parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
    if not data or not data.get("region_matrix"):
        raise HTTPException(404, f"No data for account '{account_name}'")

    if format == "csv":
        content = generate_dashboard_csv(data, account_name, provider)
        return Response(content=content, media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="dashboard-{account_name}-{provider}.csv"'})
    else:
        content = generate_dashboard_pdf(data, account_name, provider)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="dashboard-{account_name}-{provider}.pdf"'})


@app.get("/api/export/audit/{account_name}")
def export_audit(account_name: str, format: str = "pdf",
                 user: dict = Depends(get_current_user)):
    """Export audit report as PDF or CSV."""
    # Auto-detect provider
    detected_provider = "aws"
    for pid in registry.provider_ids:
        if (ACCOUNT_DATA_DIR / pid / account_name).exists():
            detected_provider = pid
            break

    plugin = registry.get(detected_provider)
    acct_dir = ACCOUNT_DATA_DIR / detected_provider / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        raise HTTPException(404, f"No data for account '{account_name}'")

    findings = plugin.auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
    severity_counts = {}
    for f in findings:
        s = f.get("severity", "INFO")
        severity_counts[s] = severity_counts.get(s, 0) + 1
    audit_data = {
        "account": account_name, "provider": detected_provider,
        "total": len(findings), "findings": findings,
        "summary": {s: severity_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]},
    }

    if format == "csv":
        content = generate_audit_csv(audit_data, account_name, detected_provider)
        return Response(content=content, media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="audit-{account_name}.csv"'})
    else:
        content = generate_audit_pdf(audit_data, account_name, detected_provider)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="audit-{account_name}.pdf"'})


# ── THREAT DETECTION ─────────────────────────────────────────────

def _detect_threats(account_name: str) -> dict:
    """Analyze cloud data for real-time threats and anomalies."""
    from providers.aws.parser import _read_json, _get_regions

    threats = []
    threat_id = 0

    # Find account data dir
    acct_dir = ACCOUNT_DATA_DIR / "aws" / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        return {"threats": [], "summary": {}, "risk_score": 0}

    regions = _get_regions(acct_dir)

    for region in regions:
        region_dir = acct_dir / region

        # ── Threat: Open ports to internet ──
        sgs = _read_json(region_dir / "ec2-describe-security-groups.json")
        for sg in sgs.get("SecurityGroups", []):
            for rule in sg.get("IpPermissions", []):
                for ipr in rule.get("IpRanges", []) + rule.get("Ipv6Ranges", []):
                    cidr = ipr.get("CidrIp", ipr.get("CidrIpv6", ""))
                    if cidr in ("0.0.0.0/0", "::/0"):
                        proto = rule.get("IpProtocol", "")
                        from_port = rule.get("FromPort")
                        to_port = rule.get("ToPort")
                        dangerous_ports = {22: "SSH", 3389: "RDP", 3306: "MySQL", 5432: "PostgreSQL", 1433: "MSSQL", 27017: "MongoDB", 6379: "Redis"}
                        port_label = dangerous_ports.get(from_port, "")

                        if proto == "-1" or port_label:
                            severity = "CRITICAL"
                            category = "network_exposure"
                            title = f"{'All traffic' if proto == '-1' else port_label + ' (port ' + str(from_port) + ')'} open to internet"
                        else:
                            severity = "HIGH"
                            category = "network_exposure"
                            title = f"Port {from_port}-{to_port} open to internet"

                        threat_id += 1
                        threats.append({
                            "id": f"TH-{threat_id:04d}",
                            "severity": severity,
                            "category": category,
                            "title": title,
                            "description": f"Security group {sg.get('GroupName', '')} ({sg.get('GroupId', '')}) allows {cidr} ingress",
                            "resource": f"{sg.get('GroupId', '')} ({sg.get('GroupName', '')})",
                            "resource_type": "SecurityGroup",
                            "region": region,
                            "detected_at": datetime.now().isoformat(),
                            "status": "active",
                            "mitre_tactic": "Initial Access",
                            "mitre_technique": "T1190 - Exploit Public-Facing Application",
                            "remediation": f"Restrict {sg.get('GroupId', '')} ingress to specific CIDRs. Remove 0.0.0.0/0 rule.",
                        })

        # ── Threat: Public EC2 instances ──
        ec2 = _read_json(region_dir / "ec2-describe-instances.json")
        for res in ec2.get("Reservations", []):
            for inst in res.get("Instances", []):
                if inst.get("PublicIpAddress") and inst.get("State", {}).get("Name") == "running":
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}",
                        "severity": "HIGH",
                        "category": "public_exposure",
                        "title": f"EC2 instance with public IP: {inst.get('PublicIpAddress')}",
                        "description": f"Instance {inst.get('InstanceId', '')} has public IP {inst.get('PublicIpAddress')} and is running",
                        "resource": inst.get("InstanceId", ""),
                        "resource_type": "EC2",
                        "region": region,
                        "detected_at": datetime.now().isoformat(),
                        "status": "active",
                        "mitre_tactic": "Initial Access",
                        "mitre_technique": "T1133 - External Remote Services",
                        "remediation": "Remove public IP. Use NAT Gateway, VPN, or Session Manager for access.",
                    })

        # ── Threat: Public RDS ──
        rds = _read_json(region_dir / "rds-describe-db-instances.json")
        for db in rds.get("DBInstances", []):
            if db.get("PubliclyAccessible"):
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}",
                    "severity": "CRITICAL",
                    "category": "data_exposure",
                    "title": f"Database publicly accessible: {db.get('DBInstanceIdentifier', '')}",
                    "description": f"RDS instance {db.get('DBInstanceIdentifier', '')} ({db.get('Engine', '')}) is publicly accessible",
                    "resource": db.get("DBInstanceIdentifier", ""),
                    "resource_type": "RDS",
                    "region": region,
                    "detected_at": datetime.now().isoformat(),
                    "status": "active",
                    "mitre_tactic": "Collection",
                    "mitre_technique": "T1530 - Data from Cloud Storage",
                    "remediation": "Disable public access on the RDS instance. Move to private subnet.",
                })

        # ── Threat: GuardDuty not enabled ──
        gd = _read_json(region_dir / "guardduty-list-detectors.json")
        if not gd.get("DetectorIds"):
            # Only flag for regions with resources
            has_resources = bool(ec2.get("Reservations") or rds.get("DBInstances"))
            if has_resources:
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}",
                    "severity": "MEDIUM",
                    "category": "detection_gap",
                    "title": f"GuardDuty not enabled in {region}",
                    "description": f"AWS GuardDuty threat detection is not active in {region} where resources exist",
                    "resource": region,
                    "resource_type": "GuardDuty",
                    "region": region,
                    "detected_at": datetime.now().isoformat(),
                    "status": "active",
                    "mitre_tactic": "Defense Evasion",
                    "mitre_technique": "T1562 - Impair Defenses",
                    "remediation": "Enable GuardDuty in this region for automated threat detection.",
                })

    # ── Threat: IAM issues ──
    for region in regions[:1]:  # IAM is global, check once
        region_dir = acct_dir / region
        iam_auth = _read_json(region_dir / "iam-get-account-authorization-details.json")
        iam_summary = _read_json(region_dir / "iam-get-account-summary.json").get("SummaryMap", {})

        if not iam_summary.get("AccountMFAEnabled"):
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}",
                "severity": "CRITICAL",
                "category": "identity_threat",
                "title": "Root account MFA not enabled",
                "description": "The AWS root account does not have MFA enabled, making it vulnerable to credential theft",
                "resource": "root-account",
                "resource_type": "IAM",
                "region": "global",
                "detected_at": datetime.now().isoformat(),
                "status": "active",
                "mitre_tactic": "Credential Access",
                "mitre_technique": "T1078 - Valid Accounts",
                "remediation": "Enable MFA on root account immediately. Use hardware MFA key.",
            })

        if iam_summary.get("AccountAccessKeysPresent"):
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}",
                "severity": "CRITICAL",
                "category": "identity_threat",
                "title": "Root account has access keys",
                "description": "Active access keys exist for the root account, posing severe security risk",
                "resource": "root-account",
                "resource_type": "IAM",
                "region": "global",
                "detected_at": datetime.now().isoformat(),
                "status": "active",
                "mitre_tactic": "Credential Access",
                "mitre_technique": "T1528 - Steal Application Access Token",
                "remediation": "Delete root access keys. Use IAM roles instead.",
            })

        users_no_mfa = [u["UserName"] for u in iam_auth.get("UserDetailList", []) if not u.get("MFADevices")]
        if users_no_mfa:
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}",
                "severity": "HIGH",
                "category": "identity_threat",
                "title": f"{len(users_no_mfa)} IAM users without MFA",
                "description": f"Users without MFA: {', '.join(users_no_mfa[:10])}{'...' if len(users_no_mfa) > 10 else ''}",
                "resource": f"{len(users_no_mfa)} users",
                "resource_type": "IAM",
                "region": "global",
                "detected_at": datetime.now().isoformat(),
                "status": "active",
                "mitre_tactic": "Credential Access",
                "mitre_technique": "T1078 - Valid Accounts",
                "remediation": "Enable MFA for all IAM users with console access.",
            })

    # ── Threat: No CloudTrail ──
    has_cloudtrail = False
    for region in regions:
        ct = _read_json(acct_dir / region / "cloudtrail-describe-trails.json")
        if ct.get("trailList"):
            has_cloudtrail = True
            break
    if not has_cloudtrail:
        threat_id += 1
        threats.append({
            "id": f"TH-{threat_id:04d}",
            "severity": "HIGH",
            "category": "detection_gap",
            "title": "No CloudTrail logging detected",
            "description": "CloudTrail is not enabled, meaning API activity is not being logged for audit purposes",
            "resource": "account",
            "resource_type": "CloudTrail",
            "region": "global",
            "detected_at": datetime.now().isoformat(),
            "status": "active",
            "mitre_tactic": "Defense Evasion",
            "mitre_technique": "T1562.008 - Disable Cloud Logs",
            "remediation": "Enable CloudTrail in all regions with log file validation.",
        })

    # ── Advanced: Secret Detection in Lambda env vars ──
    for region in regions:
        lambdas = _read_json(acct_dir / region / "lambda-list-functions.json")
        for fn in lambdas.get("Functions", []):
            env_vars = fn.get("Environment", {}).get("Variables", {})
            secret_patterns = ["KEY", "SECRET", "PASSWORD", "TOKEN", "CREDENTIAL", "API_KEY", "AUTH", "PRIVATE"]
            for key, val in env_vars.items():
                if any(p in key.upper() for p in secret_patterns) and len(val) > 8:
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}", "severity": "CRITICAL",
                        "category": "secret_exposure",
                        "title": f"Potential secret in Lambda env var: {key}",
                        "description": f"Lambda function {fn.get('FunctionName', '')} has environment variable '{key}' that may contain a secret/credential",
                        "resource": fn.get("FunctionName", ""), "resource_type": "Lambda",
                        "region": region, "detected_at": datetime.now().isoformat(), "status": "active",
                        "mitre_tactic": "Credential Access",
                        "mitre_technique": "T1552.001 - Credentials In Files",
                        "remediation": f"Move secret '{key}' to AWS Secrets Manager or Parameter Store. Reference it at runtime instead of embedding.",
                    })

    # ── Advanced: S3 Data Exposure ──
    for region in regions[:1]:  # S3 is global
        s3 = _read_json(acct_dir / region / "s3-list-buckets.json")
        bucket_count = len(s3.get("Buckets", []))
        if bucket_count > 0:
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}", "severity": "MEDIUM",
                "category": "data_exposure",
                "title": f"{bucket_count} S3 buckets require public access review",
                "description": f"Review all {bucket_count} S3 buckets for public access settings. Enable S3 Block Public Access at account level.",
                "resource": f"{bucket_count} buckets", "resource_type": "S3",
                "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                "mitre_tactic": "Collection",
                "mitre_technique": "T1530 - Data from Cloud Storage",
                "remediation": "Enable S3 Block Public Access at account level: aws s3control put-public-access-block --account-id <ID> --public-access-block-configuration BlockPublicAcls=true,IgnorePublicAcls=true,BlockPublicPolicy=true,RestrictPublicBuckets=true",
            })

    # ── Advanced: Encryption Audit ──
    for region in regions:
        # Unencrypted snapshots
        snaps = _read_json(acct_dir / region / "ec2-describe-snapshots.json")
        unenc_snaps = [s for s in snaps.get("Snapshots", []) if not s.get("Encrypted")]
        if unenc_snaps:
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}", "severity": "HIGH",
                "category": "encryption_gap",
                "title": f"{len(unenc_snaps)} unencrypted EBS snapshots in {region}",
                "description": f"Found {len(unenc_snaps)} snapshots without encryption: {', '.join(s.get('SnapshotId','') for s in unenc_snaps[:5])}",
                "resource": f"{len(unenc_snaps)} snapshots", "resource_type": "Snapshot",
                "region": region, "detected_at": datetime.now().isoformat(), "status": "active",
                "mitre_tactic": "Collection",
                "mitre_technique": "T1530 - Data from Cloud Storage",
                "remediation": "Copy snapshots with encryption enabled. Enable EBS encryption by default for the account.",
            })

        # Unencrypted RDS
        rds = _read_json(acct_dir / region / "rds-describe-db-instances.json")
        for db in rds.get("DBInstances", []):
            if not db.get("StorageEncrypted"):
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}", "severity": "HIGH",
                    "category": "encryption_gap",
                    "title": f"RDS instance without encryption: {db.get('DBInstanceIdentifier', '')}",
                    "description": f"Database {db.get('DBInstanceIdentifier', '')} ({db.get('Engine', '')}) does not have storage encryption enabled",
                    "resource": db.get("DBInstanceIdentifier", ""), "resource_type": "RDS",
                    "region": region, "detected_at": datetime.now().isoformat(), "status": "active",
                    "mitre_tactic": "Collection",
                    "mitre_technique": "T1565 - Data Manipulation",
                    "remediation": "Encrypt RDS: Create encrypted snapshot → Restore from snapshot → Switch DNS/endpoint.",
                })

    # ── Advanced: Lateral Movement - Overly permissive IAM ──
    for region in regions[:1]:
        iam_auth = _read_json(acct_dir / region / "iam-get-account-authorization-details.json")
        for role in iam_auth.get("RoleDetailList", []):
            trust = role.get("AssumeRolePolicyDocument", {})
            if isinstance(trust, str):
                try:
                    trust = json.loads(trust)
                except Exception:
                    continue
            for stmt in trust.get("Statement", []):
                principal = stmt.get("Principal", {})
                if isinstance(principal, str) and principal == "*":
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}", "severity": "CRITICAL",
                        "category": "lateral_movement",
                        "title": f"IAM role assumable by anyone: {role.get('RoleName', '')}",
                        "description": f"Role {role.get('RoleName', '')} has trust policy with Principal: '*', allowing any AWS account to assume it",
                        "resource": role.get("RoleName", ""), "resource_type": "IAM Role",
                        "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                        "mitre_tactic": "Lateral Movement",
                        "mitre_technique": "T1550.001 - Application Access Token",
                        "remediation": f"Restrict trust policy for role {role.get('RoleName', '')} to specific account IDs or services.",
                    })

            # Check for admin policies
            for policy in role.get("AttachedManagedPolicies", []):
                if "AdministratorAccess" in policy.get("PolicyName", ""):
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}", "severity": "HIGH",
                        "category": "lateral_movement",
                        "title": f"Role with full admin access: {role.get('RoleName', '')}",
                        "description": f"Role {role.get('RoleName', '')} has AdministratorAccess policy attached — full control over all resources",
                        "resource": role.get("RoleName", ""), "resource_type": "IAM Role",
                        "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                        "mitre_tactic": "Privilege Escalation",
                        "mitre_technique": "T1078.004 - Cloud Accounts",
                        "remediation": "Apply least-privilege principle. Replace AdministratorAccess with specific service policies.",
                    })

    # ── Advanced: Compliance Drift ──
    try:
        latest_compliance = _compliance_store.get_latest(account_name)
        history = _compliance_store.get_history(account_name)
        if latest_compliance and len(history) >= 2:
            current_score = latest_compliance.get("summary", {}).get("overall_score", 0)
            prev_score = history[1].get("overall_score", 0)  # index 0 is latest, 1 is previous
            if current_score < prev_score:
                drift = prev_score - current_score
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}", "severity": "HIGH" if drift > 10 else "MEDIUM",
                    "category": "compliance_drift",
                    "title": f"Compliance score dropped {drift}% (from {prev_score}% to {current_score}%)",
                    "description": f"Overall compliance score decreased by {drift} percentage points since the previous scan. This indicates new misconfigurations or policy violations.",
                    "resource": "compliance", "resource_type": "Compliance",
                    "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                    "mitre_tactic": "Defense Evasion",
                    "mitre_technique": "T1562 - Impair Defenses",
                    "remediation": f"Review compliance results at /compliance. Focus on newly failed checks to restore score from {current_score}% to {prev_score}%+.",
                })
    except Exception:
        pass

    # ── Build Attack Paths ──
    attack_paths = []

    # Path 1: Internet → Open Port → EC2 → IAM Role → Data
    public_instances = [t for t in threats if t["category"] == "public_exposure"]
    open_ports = [t for t in threats if t["category"] == "network_exposure"]
    data_exposure = [t for t in threats if t["category"] == "data_exposure"]
    identity_threats = [t for t in threats if t["category"] == "identity_threat"]

    if open_ports and public_instances:
        attack_paths.append({
            "id": "AP-001",
            "name": "Internet → EC2 via Open Ports",
            "severity": "CRITICAL",
            "description": "Attacker can reach EC2 instances through open security group rules and public IPs",
            "steps": [
                {"step": 1, "action": "Scan public IPs", "detail": f"{len(public_instances)} instance(s) with public IPs", "type": "recon"},
                {"step": 2, "action": "Exploit open port", "detail": f"{len(open_ports)} open port rule(s) to 0.0.0.0/0", "type": "exploit"},
                {"step": 3, "action": "Gain instance access", "detail": "SSH/RDP access via open ports", "type": "access"},
                {"step": 4, "action": "Steal IAM credentials", "detail": "Query instance metadata for IAM role credentials", "type": "escalate"},
            ],
            "impact": "Full instance compromise, potential lateral movement via IAM role",
            "mitre_chain": ["T1595 - Active Scanning", "T1190 - Exploit Public-Facing App", "T1078 - Valid Accounts", "T1552 - Unsecured Credentials"],
        })

    if data_exposure:
        attack_paths.append({
            "id": "AP-002",
            "name": "Internet → Public Database → Data Exfiltration",
            "severity": "CRITICAL",
            "description": "Publicly accessible databases can be directly attacked from the internet",
            "steps": [
                {"step": 1, "action": "Discover public endpoint", "detail": f"{len(data_exposure)} public database(s)", "type": "recon"},
                {"step": 2, "action": "Brute force credentials", "detail": "Default or weak database credentials", "type": "exploit"},
                {"step": 3, "action": "Exfiltrate data", "detail": "Direct SQL access to all data", "type": "exfiltrate"},
            ],
            "impact": "Complete data breach — customer data, PII, financial records",
            "mitre_chain": ["T1595 - Active Scanning", "T1110 - Brute Force", "T1530 - Data from Cloud Storage"],
        })

    if identity_threats:
        attack_paths.append({
            "id": "AP-003",
            "name": "Credential Theft → Account Takeover",
            "severity": "HIGH",
            "description": "Weak identity controls allow account compromise through credential theft",
            "steps": [
                {"step": 1, "action": "Phish IAM user", "detail": f"{len([t for t in identity_threats if 'without MFA' in t.get('title','')])} users without MFA protection", "type": "recon"},
                {"step": 2, "action": "Use stolen credentials", "detail": "No MFA = single-factor compromise", "type": "exploit"},
                {"step": 3, "action": "Escalate privileges", "detail": "Access cloud console or API with stolen identity", "type": "escalate"},
                {"step": 4, "action": "Persistent access", "detail": "Create backdoor IAM users or access keys", "type": "persist"},
            ],
            "impact": "Full account takeover, resource manipulation, data theft",
            "mitre_chain": ["T1566 - Phishing", "T1078 - Valid Accounts", "T1098 - Account Manipulation"],
        })

    secret_threats = [t for t in threats if t["category"] == "secret_exposure"]
    if secret_threats:
        attack_paths.append({
            "id": "AP-004",
            "name": "Exposed Secrets → Service Compromise",
            "severity": "CRITICAL",
            "description": "Hardcoded secrets in Lambda functions can be extracted and used to access other services",
            "steps": [
                {"step": 1, "action": "Access Lambda config", "detail": f"{len(secret_threats)} functions with embedded secrets", "type": "recon"},
                {"step": 2, "action": "Extract credentials", "detail": "Read environment variables via console or API", "type": "exploit"},
                {"step": 3, "action": "Access target service", "detail": "Use extracted API keys/tokens for lateral movement", "type": "escalate"},
            ],
            "impact": "Compromise of connected services — databases, APIs, third-party systems",
            "mitre_chain": ["T1552.001 - Credentials In Files", "T1550 - Use Alternate Auth Material"],
        })

    lateral = [t for t in threats if t["category"] == "lateral_movement"]
    encryption = [t for t in threats if t["category"] == "encryption_gap"]

    if lateral:
        attack_paths.append({
            "id": "AP-005",
            "name": "Over-Privileged Role → Cross-Account Pivot",
            "severity": "HIGH",
            "description": "Overly permissive IAM roles can be assumed to pivot across the infrastructure",
            "steps": [
                {"step": 1, "action": "Discover assumable roles", "detail": f"{len(lateral)} roles with excessive permissions", "type": "recon"},
                {"step": 2, "action": "Assume admin role", "detail": "Use sts:AssumeRole with wildcard trust", "type": "escalate"},
                {"step": 3, "action": "Full infrastructure access", "detail": "Admin role grants control over all resources", "type": "access"},
            ],
            "impact": "Complete infrastructure compromise — create, modify, delete any resource",
            "mitre_chain": ["T1078.004 - Cloud Accounts", "T1550.001 - Application Access Token"],
        })

    # Sort by severity
    sev_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    threats.sort(key=lambda t: sev_order.get(t["severity"], 5))

    # Summary
    sev_counts = {}
    cat_counts = {}
    region_counts = {}
    resource_type_counts = {}
    for t in threats:
        sev_counts[t["severity"]] = sev_counts.get(t["severity"], 0) + 1
        cat_counts[t["category"]] = cat_counts.get(t["category"], 0) + 1
        region_counts[t["region"]] = region_counts.get(t["region"], 0) + 1
        resource_type_counts[t["resource_type"]] = resource_type_counts.get(t["resource_type"], 0) + 1

    total = len(threats)
    crit = sev_counts.get("CRITICAL", 0)
    high = sev_counts.get("HIGH", 0)
    risk_score = min(100, max(0, 100 - crit * 12 - high * 4 - sev_counts.get("MEDIUM", 0) * 1))

    return {
        "account": account_name,
        "scanned_at": datetime.now().isoformat(),
        "threats": threats,
        "attack_paths": attack_paths,
        "total": total,
        "risk_score": risk_score,
        "summary": {
            "severity": {s: sev_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]},
            "categories": cat_counts,
            "regions": dict(sorted(region_counts.items(), key=lambda x: x[1], reverse=True)[:10]),
            "resource_types": resource_type_counts,
        },
    }


@app.get("/api/threats/{account_name}")
def get_threats(account_name: str, user: dict = Depends(get_current_user)):
    """Advanced threat detection with attack paths, secret scanning, encryption audit, and MITRE ATT&CK mapping."""
    result = _detect_threats(account_name)
    if not result["threats"] and result["risk_score"] == 0:
        raise HTTPException(404, f"No data for account '{account_name}'")
    return result


# ── COMPLIANCE MODULE ────────────────────────────────────────────
from compliance.engine import ComplianceScanner, ComplianceStore
from compliance.frameworks import get_all_frameworks, get_framework

_compliance_scanner = ComplianceScanner(ACCOUNT_DATA_DIR)
_compliance_store = ComplianceStore(BASE_DIR)


class ComplianceScanRequest(BaseModel):
    frameworks: list = []  # empty = all frameworks


@app.get("/api/compliance/frameworks")
def list_compliance_frameworks(user: dict = Depends(get_current_user)):
    """List all supported compliance frameworks."""
    frameworks = get_all_frameworks()
    return {"frameworks": [
        {
            "id": fw.framework_id, "name": fw.name, "version": fw.version,
            "description": fw.description, "category": fw.category,
            "icon": fw.icon, "color": fw.color,
            "total_controls": len(fw.controls),
            "total_checks": fw.total_checks,
        }
        for fw in frameworks.values()
    ]}


@app.get("/api/compliance/frameworks/{framework_id}")
def get_compliance_framework(framework_id: str, user: dict = Depends(get_current_user)):
    """Get detailed framework info with controls and checks."""
    fw = get_framework(framework_id)
    if not fw:
        raise HTTPException(404, f"Framework '{framework_id}' not found")
    return {
        "id": fw.framework_id, "name": fw.name, "version": fw.version,
        "description": fw.description, "category": fw.category,
        "icon": fw.icon, "color": fw.color,
        "controls": [
            {
                "control_id": c.control_id, "title": c.title,
                "description": c.description, "section": c.section,
                "severity": c.severity,
                "checks": [
                    {"check_id": ch.check_id, "title": ch.title, "description": ch.description,
                     "resource_type": ch.resource_type, "severity": ch.severity, "remediation": ch.remediation}
                    for ch in c.checks
                ],
            }
            for c in fw.controls
        ],
    }


@app.post("/api/compliance/scan/{account_name}")
def run_compliance_scan(account_name: str, req: ComplianceScanRequest = ComplianceScanRequest(),
                        user: dict = Depends(get_current_user)):
    """Run compliance scan against all or selected frameworks."""
    fw_ids = req.frameworks if req.frameworks else None
    results = _compliance_scanner.scan(account_name, fw_ids)
    if "error" in results:
        raise HTTPException(404, results["error"])
    _compliance_store.save(account_name, results)
    return results


@app.get("/api/compliance/results/{account_name}")
def get_compliance_results(account_name: str, user: dict = Depends(get_current_user)):
    """Get latest compliance scan results."""
    results = _compliance_store.get_latest(account_name)
    if not results:
        raise HTTPException(404, f"No compliance results for '{account_name}'. Run a scan first.")
    return results


@app.get("/api/compliance/history/{account_name}")
def get_compliance_history(account_name: str, user: dict = Depends(get_current_user)):
    """Get compliance scan history for drift tracking."""
    history = _compliance_store.get_history(account_name)
    return {"account": account_name, "history": history}


@app.get("/api/compliance/results/{account_name}/export")
def export_compliance_report(account_name: str, format: str = "pdf",
                             user: dict = Depends(get_current_user)):
    """Export compliance report as PDF, Excel, or JSON."""
    results = _compliance_store.get_latest(account_name)
    if not results:
        raise HTTPException(404, f"No compliance results for '{account_name}'")

    if format == "json":
        return Response(
            content=json.dumps(results, indent=2, default=str),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="compliance-{account_name}.json"'},
        )
    elif format == "csv":
        lines = ["Framework,Control,Check,Status,Severity,Reason,Remediation"]
        for fw_id, fw in results.get("frameworks", {}).items():
            for ctrl in fw.get("controls", []):
                for ch in ctrl.get("checks", []):
                    lines.append(','.join([
                        f'"{fw["name"]}"', f'"{ctrl["section"]}"',
                        f'"{ch["title"]}"', ch["status"], ch["severity"],
                        f'"{ch["reason"]}"', f'"{ch["remediation"][:100]}"',
                    ]))
        content = '\n'.join(lines)
        return Response(content=content, media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="compliance-{account_name}.csv"'})
    else:
        # PDF export using existing report generator patterns
        content = _generate_compliance_pdf(results)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="compliance-{account_name}.pdf"'})


def _generate_compliance_pdf(results):
    """Generate compliance PDF report."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1E293B'), spaceAfter=5)
    h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=13, textColor=colors.HexColor('#4F46E5'), spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle('B', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#475569'), leading=12)
    small = ParagraphStyle('S', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94A3B8'))

    elements = []
    summary = results.get("summary", {})
    elements.append(Paragraph("CloudSentinel Compliance Report", title_s))
    elements.append(Paragraph(f"Account: {results.get('account')} | Scanned: {results.get('scanned_at', '')[:19]} | Score: {summary.get('overall_score', 0)}%", small))
    elements.append(Spacer(1, 10))

    # Summary table
    elements.append(Paragraph("Executive Summary", h1))
    summary_data = [
        ["Overall Score", f"{summary.get('overall_score', 0)}%", "Total Checks", str(summary.get('total_checks', 0))],
        ["Passed", str(summary.get('passed', 0)), "Failed", str(summary.get('failed', 0))],
        ["Warnings", str(summary.get('warnings', 0)), "Critical", str(summary.get('severity_counts', {}).get('CRITICAL', 0))],
    ]
    t = Table(summary_data, colWidths=[80, 60, 80, 60])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('FONTNAME', (3, 0), (3, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)

    # Framework scores
    elements.append(Paragraph("Framework Compliance Scores", h1))
    fw_rows = [["Framework", "Score", "Passed", "Failed", "Status"]]
    for fw_id, fw in results.get("frameworks", {}).items():
        status = "PASS" if fw["score"] >= 80 else "WARN" if fw["score"] >= 50 else "FAIL"
        fw_rows.append([fw["name"], f"{fw['score']}%", str(fw["passed"]), str(fw["failed"]), status])
    t = Table(fw_rows, colWidths=[140, 50, 50, 50, 50])
    style_list = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')), ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ]
    sev_bg = {"PASS": '#D1FAE5', "WARN": '#FEF9C3', "FAIL": '#FEE2E2'}
    for i, row in enumerate(fw_rows[1:], 1):
        style_list.append(('BACKGROUND', (4, i), (4, i), colors.HexColor(sev_bg.get(row[4], '#F1F5F9'))))
    t.setStyle(TableStyle(style_list))
    elements.append(t)

    elements.append(PageBreak())

    # Findings
    elements.append(Paragraph("Failed Compliance Checks", h1))
    findings = results.get("all_findings", [])
    if findings:
        find_rows = [["Severity", "Check", "Reason", "Framework"]]
        for f in findings[:80]:
            fw_name = results["frameworks"].get(f.get("framework_id", ""), {}).get("name", "")
            find_rows.append([f["severity"], f["title"][:40], f["reason"][:50], fw_name[:20]])
        t = Table(find_rows, colWidths=[50, 140, 130, 80])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(t)

    # AI Recommendations
    elements.append(PageBreak())
    elements.append(Paragraph("AI Recommendations", h1))
    for rec in results.get("ai_recommendations", []):
        elements.append(Paragraph(f"<b>[{rec['priority']}] {rec['title']}</b>", body))
        elements.append(Paragraph(rec["description"], body))
        for item in rec.get("items", []):
            elements.append(Paragraph(f"• {item}", body))
        elements.append(Spacer(1, 6))

    doc.build(elements)
    return buf.getvalue()


# ── AI CHAT ──────────────────────────────────────────────────────
class AiChatRequest(BaseModel):
    message: str
    history: list = []

# Well-Architected Framework pillars for AI analysis
WAF_PILLARS = {
    "security": {
        "name": "Security",
        "checks": [
            ("Root MFA", lambda d: d.get("iam_summary", {}).get("AccountMFAEnabled", False), "Enable MFA on root account"),
            ("No root keys", lambda d: not d.get("iam_summary", {}).get("AccountAccessKeysPresent", True), "Remove root account access keys"),
            ("IAM MFA", lambda d: d.get("iam_users_no_mfa", 1) == 0, "Enable MFA for all IAM users"),
            ("No open SGs", lambda d: d.get("open_security_groups", 1) == 0, "Restrict security groups open to 0.0.0.0/0"),
            ("No public RDS", lambda d: d.get("public_summary", {}).get("rds", 1) == 0, "Disable public access on RDS instances"),
        ],
    },
    "reliability": {
        "name": "Reliability",
        "checks": [
            ("Multi-AZ", lambda d: len([r for r, s in d.get("regions", {}).items() if s.get("has_resources")]) > 1, "Deploy across multiple availability zones"),
            ("Snapshots exist", lambda d: d.get("totals", {}).get("snapshots", 0) > 0, "Create regular EBS/RDS snapshots"),
            ("ELBs present", lambda d: d.get("totals", {}).get("elbs", 0) > 0, "Use load balancers for high availability"),
        ],
    },
    "performance": {
        "name": "Performance Efficiency",
        "checks": [
            ("Right-sized instances", lambda d: True, "Review instance types for right-sizing"),
            ("Lambda usage", lambda d: d.get("totals", {}).get("lambdas", 0) > 0, "Consider serverless for variable workloads"),
        ],
    },
    "cost": {
        "name": "Cost Optimization",
        "checks": [
            ("No stopped EC2", lambda d: all(r.get("instances_stopped", 0) == 0 for r in d.get("regions", {}).values()), "Terminate or schedule stopped instances"),
            ("Snapshot cleanup", lambda d: d.get("totals", {}).get("snapshots", 0) < 50, "Review and clean up old snapshots"),
        ],
    },
    "operational": {
        "name": "Operational Excellence",
        "checks": [
            ("CloudTrail", lambda d: any(r.get("cloudtrail_trails", 0) > 0 for r in d.get("regions", {}).values()), "Enable CloudTrail in all regions"),
            ("GuardDuty", lambda d: any(r.get("guardduty_enabled", False) for r in d.get("regions", {}).values()), "Enable GuardDuty for threat detection"),
        ],
    },
    "sustainability": {
        "name": "Sustainability",
        "checks": [
            ("Region consolidation", lambda d: len([r for r, s in d.get("regions", {}).items() if s.get("has_resources")]) <= 5, "Consolidate resources to fewer regions"),
        ],
    },
}


def _analyze_waf(dashboard_data):
    """Analyze dashboard data against Well-Architected Framework pillars."""
    results = {}
    for pillar_id, pillar in WAF_PILLARS.items():
        checks = []
        passed = 0
        for check_name, check_fn, recommendation in pillar["checks"]:
            try:
                ok = check_fn(dashboard_data)
            except Exception:
                ok = False
            checks.append({"name": check_name, "passed": ok, "recommendation": recommendation})
            if ok:
                passed += 1
        total = len(checks)
        results[pillar_id] = {
            "name": pillar["name"],
            "score": round(passed / total * 100) if total > 0 else 0,
            "passed": passed,
            "total": total,
            "checks": checks,
        }
    overall = sum(p["score"] for p in results.values()) / len(results) if results else 0
    return {"overall_score": round(overall), "pillars": results}


def _generate_ai_response(message, dashboard_data, audit_findings, waf_analysis):
    """Generate an AI-style response based on actual cloud data."""
    lower = message.lower()
    lines = []

    if any(w in lower for w in ["well-architected", "waf", "pillar", "framework", "architect"]):
        lines.append("## AWS Well-Architected Framework Analysis\n")
        for pid, pillar in waf_analysis.get("pillars", {}).items():
            icon = "✅" if pillar["score"] >= 80 else "⚠️" if pillar["score"] >= 50 else "❌"
            lines.append(f"{icon} **{pillar['name']}**: {pillar['score']}% ({pillar['passed']}/{pillar['total']} checks passed)")
            for check in pillar["checks"]:
                status = "✓" if check["passed"] else "✗"
                lines.append(f"   {status} {check['name']}" + ("" if check["passed"] else f" — {check['recommendation']}"))
        lines.append(f"\n**Overall Score: {waf_analysis.get('overall_score', 0)}%**")
        return "\n".join(lines)

    if any(w in lower for w in ["risk", "top", "critical", "threat", "issue"]):
        critical = [f for f in audit_findings if f.get("severity") == "CRITICAL"]
        high = [f for f in audit_findings if f.get("severity") == "HIGH"]
        lines.append("## Top Security Risks\n")
        if critical:
            lines.append(f"**🔴 {len(critical)} Critical Finding(s):**")
            for f in critical[:5]:
                lines.append(f"- {f.get('title', f.get('issue', 'Unknown'))}")
                if f.get('resource'):
                    lines.append(f"  Resource: `{f['resource']}`")
        if high:
            lines.append(f"\n**🟠 {len(high)} High Finding(s):**")
            for f in high[:5]:
                lines.append(f"- {f.get('title', f.get('issue', 'Unknown'))}")
        if not critical and not high:
            lines.append("No critical or high severity findings detected. Your security posture looks good!")
        lines.append("\n**Recommendation:** Address critical findings immediately, then work through high-severity items.")
        return "\n".join(lines)

    if any(w in lower for w in ["score", "improve", "better", "increase"]):
        score = dashboard_data.get("security_score", 0)
        lines.append(f"## Security Score: {score}/100\n")
        waf = waf_analysis.get("pillars", {})
        weak = sorted(waf.items(), key=lambda x: x[1]["score"])
        lines.append("**Areas to improve (sorted by priority):**\n")
        for pid, pillar in weak:
            if pillar["score"] < 100:
                failed = [c for c in pillar["checks"] if not c["passed"]]
                for c in failed:
                    lines.append(f"- **{pillar['name']}**: {c['recommendation']}")
        return "\n".join(lines)

    if any(w in lower for w in ["posture", "summary", "overview", "status"]):
        t = dashboard_data.get("totals", {})
        lines.append("## Cloud Security Posture Summary\n")
        lines.append(f"- **EC2 Instances:** {t.get('instances', 0)}")
        lines.append(f"- **S3 Buckets:** {t.get('buckets', 0)}")
        lines.append(f"- **Lambda Functions:** {t.get('lambdas', 0)}")
        lines.append(f"- **RDS Instances:** {t.get('rds', 0)}")
        lines.append(f"- **Security Groups:** {t.get('security_groups', 0)}")
        lines.append(f"- **Open Security Groups:** {dashboard_data.get('open_security_groups', 0)}")
        lines.append(f"- **Public IPs:** {len(dashboard_data.get('public_ips', []))}")
        lines.append(f"- **Security Score:** {dashboard_data.get('security_score', 0)}/100")
        lines.append(f"- **Regions Scanned:** {dashboard_data.get('regions_scanned', 0)}")
        sev = {}
        for f in audit_findings:
            s = f.get("severity", "INFO")
            sev[s] = sev.get(s, 0) + 1
        lines.append(f"\n**Audit Findings:** {len(audit_findings)} total")
        for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]:
            if sev.get(s, 0) > 0:
                lines.append(f"  - {s}: {sev[s]}")
        return "\n".join(lines)

    if any(w in lower for w in ["public", "exposed", "internet"]):
        pips = dashboard_data.get("public_ips", [])
        ps = dashboard_data.get("public_summary", {})
        lines.append(f"## Publicly Exposed Resources ({len(pips)} total)\n")
        if ps.get("ec2", 0) > 0:
            lines.append(f"- **{ps['ec2']} EC2 instance(s)** with public IPs")
        if ps.get("rds", 0) > 0:
            lines.append(f"- **{ps['rds']} RDS instance(s)** publicly accessible ⚠️")
        if ps.get("elb", 0) > 0:
            lines.append(f"- **{ps['elb']} Load Balancer(s)** internet-facing")
        osg = dashboard_data.get("open_security_groups", 0)
        if osg > 0:
            lines.append(f"- **{osg} Security Group rule(s)** open to 0.0.0.0/0")
        if not pips:
            lines.append("No publicly exposed resources detected. Great job!")
        else:
            lines.append("\n**Recommendations:**")
            lines.append("- Move RDS instances to private subnets")
            lines.append("- Use VPC endpoints instead of public access")
            lines.append("- Restrict security group ingress to known CIDRs")
        return "\n".join(lines)

    if any(w in lower for w in ["mfa", "authentication", "iam", "user"]):
        iam = dashboard_data.get("iam_summary", {})
        lines.append("## IAM & Authentication Status\n")
        lines.append(f"- **Root MFA:** {'✅ Enabled' if iam.get('AccountMFAEnabled') else '❌ Disabled'}")
        lines.append(f"- **Root Access Keys:** {'❌ Present (remove!)' if iam.get('AccountAccessKeysPresent') else '✅ None'}")
        lines.append(f"- **IAM Users:** {iam.get('Users', 0)}")
        lines.append(f"- **Users without MFA:** {dashboard_data.get('iam_users_no_mfa', 0)}")
        lines.append(f"- **IAM Roles:** {iam.get('Roles', 0)}")
        lines.append(f"- **Policies:** {iam.get('Policies', 0)}")
        return "\n".join(lines)

    if any(w in lower for w in ["compliance", "regulation", "gdpr", "soc", "cis", "benchmark"]):
        lines.append("## Compliance Assessment\n")
        waf = waf_analysis.get("pillars", {})
        sec = waf.get("security", {})
        ops = waf.get("operational", {})
        lines.append(f"**CIS Benchmark Alignment:** {sec.get('score', 0)}%")
        lines.append(f"**Operational Readiness:** {ops.get('score', 0)}%\n")
        lines.append("**Key Compliance Controls:**")
        lines.append(f"- Encryption at rest: Review S3 bucket policies")
        lines.append(f"- Audit logging: {'✅' if ops.get('score', 0) > 50 else '⚠️'} CloudTrail status")
        lines.append(f"- Access control: {'✅' if sec.get('score', 0) > 60 else '⚠️'} IAM policies")
        lines.append(f"- Network security: {dashboard_data.get('open_security_groups', 0)} open rules to review")
        return "\n".join(lines)

    if any(w in lower for w in ["recommend", "suggest", "best practice", "action", "fix"]):
        lines.append("## Actionable Recommendations\n")
        lines.append("**🔴 Immediate (Critical):**")
        waf = waf_analysis.get("pillars", {})
        for pid, pillar in waf.items():
            for c in pillar.get("checks", []):
                if not c["passed"]:
                    lines.append(f"- [{pillar['name']}] {c['recommendation']}")
        lines.append("\n**🟡 Short-term:**")
        lines.append("- Review and tag all resources for cost tracking")
        lines.append("- Set up AWS Config for continuous compliance")
        lines.append("- Enable VPC Flow Logs for network monitoring")
        lines.append("\n**🟢 Long-term:**")
        lines.append("- Implement Infrastructure as Code (Terraform/CloudFormation)")
        lines.append("- Set up automated remediation with Lambda")
        lines.append("- Establish a cloud security governance framework")
        return "\n".join(lines)

    # Default response
    lines.append("I can help you analyze your cloud infrastructure. Try asking about:")
    lines.append("")
    lines.append("- **'Well-Architected analysis'** — Check against AWS WAF pillars")
    lines.append("- **'Top security risks'** — Critical and high-severity findings")
    lines.append("- **'Security score'** — How to improve your score")
    lines.append("- **'Cloud posture summary'** — Resource and security overview")
    lines.append("- **'Publicly exposed resources'** — Internet-facing services")
    lines.append("- **'IAM status'** — Users, MFA, and access analysis")
    lines.append("- **'Compliance assessment'** — CIS, SOC2, GDPR alignment")
    lines.append("- **'Recommendations'** — Actionable improvement steps")
    return "\n".join(lines)


@app.post("/api/ai/chat")
def ai_chat(req: AiChatRequest, user: dict = Depends(get_current_user)):
    """AI-powered security analysis chat."""
    # Gather data from all accounts
    dashboard_data = {}
    audit_findings = []

    for pid in registry.provider_ids:
        plugin = registry.get(pid)
        provider_dir = ACCOUNT_DATA_DIR / pid
        if not provider_dir.exists():
            continue
        for acct_dir in provider_dir.iterdir():
            if not acct_dir.is_dir():
                continue
            name = acct_dir.name
            try:
                dash = plugin.parser.parse_dashboard(name, ACCOUNT_DATA_DIR)
                if dash and dash.get("totals"):
                    dashboard_data = dash  # Use the most recent
                findings = plugin.auditor.run_audit(name, ACCOUNT_DATA_DIR)
                audit_findings.extend(findings)
            except Exception:
                pass

    # Also check legacy dirs
    if not dashboard_data:
        for acct_dir in ACCOUNT_DATA_DIR.iterdir():
            if acct_dir.is_dir() and not acct_dir.name in registry.provider_ids:
                try:
                    dash = registry.get("aws").parser.parse_dashboard(acct_dir.name, ACCOUNT_DATA_DIR)
                    if dash and dash.get("totals"):
                        dashboard_data = dash
                    findings = registry.get("aws").auditor.run_audit(acct_dir.name, ACCOUNT_DATA_DIR)
                    audit_findings.extend(findings)
                except Exception:
                    pass

    waf_analysis = _analyze_waf(dashboard_data) if dashboard_data else {"overall_score": 0, "pillars": {}}

    response = _generate_ai_response(req.message, dashboard_data, audit_findings, waf_analysis)
    return {"response": response, "waf_score": waf_analysis.get("overall_score", 0)}


# ── WELL-ARCHITECTED REPORT ──────────────────────────────────────
@app.get("/api/waf/{account_name}")
def get_waf_report(account_name: str, user: dict = Depends(get_current_user)):
    """Get Well-Architected Framework analysis for an account."""
    dashboard_data = {}
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            try:
                dashboard_data = registry.get(pid).parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
                break
            except Exception:
                pass
    if not dashboard_data:
        legacy_dir = ACCOUNT_DATA_DIR / account_name
        if legacy_dir.exists():
            try:
                dashboard_data = registry.get("aws").parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass

    if not dashboard_data:
        raise HTTPException(404, f"No data for account '{account_name}'")

    return _analyze_waf(dashboard_data)


# ── COMPREHENSIVE REPORT ─────────────────────────────────────────

def _gather_account_data(account_name: str):
    """Gather all data for an account across all providers."""
    dashboard_data = {}
    audit_findings = []
    iam_data = {}
    sg_data = {}
    resources_data = {}
    detected_provider = "aws"

    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            detected_provider = pid
            plugin = registry.get(pid)
            try:
                dashboard_data = plugin.parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass
            try:
                audit_findings = plugin.auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass
            try:
                resources_data = plugin.parser.parse_resources(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass
            break

    # IAM (AWS-specific)
    try:
        from providers.aws.parser import _read_json, _get_regions
        idir = ACCOUNT_DATA_DIR / "aws" / account_name
        if not idir.exists():
            idir = ACCOUNT_DATA_DIR / account_name
        regions = _get_regions(idir)
        if regions:
            rdir = idir / regions[0]
            auth = _read_json(rdir / "iam-get-account-authorization-details.json")
            summary = _read_json(rdir / "iam-get-account-summary.json")
            iam_data = {
                "summary": summary.get("SummaryMap", {}),
                "users": [{"name": u.get("UserName"), "arn": u.get("Arn"),
                           "mfa_devices": u.get("MFADevices", []),
                           "policies": [p["PolicyName"] for p in u.get("AttachedManagedPolicies", [])],
                           "groups": u.get("GroupList", [])}
                          for u in auth.get("UserDetailList", [])],
                "roles": [{"name": r.get("RoleName"), "arn": r.get("Arn"),
                           "policies": [p["PolicyName"] for p in r.get("AttachedManagedPolicies", [])]}
                          for r in auth.get("RoleDetailList", [])],
            }
    except Exception:
        pass

    # Security groups
    try:
        from providers.aws.parser import _read_json, _get_regions
        sgdir = ACCOUNT_DATA_DIR / "aws" / account_name
        if not sgdir.exists():
            sgdir = ACCOUNT_DATA_DIR / account_name
        regions = _get_regions(sgdir)
        risky = []
        for region in regions:
            sgs = _read_json(sgdir / region / "ec2-describe-security-groups.json")
            for sg in sgs.get("SecurityGroups", []):
                open_rules = []
                for rule in sg.get("IpPermissions", []):
                    for ipr in rule.get("IpRanges", []):
                        if ipr.get("CidrIp") == "0.0.0.0/0":
                            open_rules.append({"protocol": rule.get("IpProtocol"), "from_port": rule.get("FromPort"), "to_port": rule.get("ToPort")})
                if open_rules:
                    risky.append({"region": region, "group_id": sg["GroupId"], "group_name": sg["GroupName"],
                                  "vpc_id": sg.get("VpcId"), "open_rules": open_rules,
                                  "severity": "CRITICAL" if any(r.get("from_port") in (22, 3389, 3306, 5432) or r.get("protocol") == "-1" for r in open_rules) else "HIGH"})
        sg_data = {"risky_groups": risky}
    except Exception:
        pass

    waf = _analyze_waf(dashboard_data) if dashboard_data else {"overall_score": 0, "pillars": {}}

    # Build AI recommendations
    ai_sections = {}
    for topic in ["summary", "risk", "recommend", "compliance", "public", "iam"]:
        ai_sections[topic] = _generate_ai_response(topic, dashboard_data, audit_findings, waf)

    # Severity counts
    sev_counts = {}
    for f in audit_findings:
        s = f.get("severity", "INFO")
        sev_counts[s] = sev_counts.get(s, 0) + 1

    return {
        "account": account_name,
        "provider": detected_provider,
        "generated_at": datetime.now().isoformat(),
        "dashboard": dashboard_data,
        "audit": {
            "total": len(audit_findings),
            "findings": audit_findings,
            "summary": {s: sev_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]},
        },
        "iam": iam_data,
        "security_groups": sg_data,
        "waf": waf,
        "ai_recommendations": ai_sections,
    }


@app.get("/api/report/{account_name}")
def get_comprehensive_report(account_name: str, user: dict = Depends(get_current_user)):
    """Get combined report data for all features."""
    data = _gather_account_data(account_name)
    if not data["dashboard"]:
        raise HTTPException(404, f"No data for account '{account_name}'")
    return data


@app.get("/api/report/{account_name}/export")
def export_comprehensive_report(account_name: str, format: str = "pdf",
                                 user: dict = Depends(get_current_user)):
    """Export comprehensive report as PDF or Excel."""
    data = _gather_account_data(account_name)
    if not data["dashboard"]:
        raise HTTPException(404, f"No data for account '{account_name}'")

    if format == "excel":
        content = _generate_excel_report(data)
        return Response(content=content,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="CloudSentinel-Report-{account_name}-{datetime.now().strftime("%Y%m%d")}.xlsx"'})
    else:
        content = _generate_comprehensive_pdf(data)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="CloudSentinel-Report-{account_name}-{datetime.now().strftime("%Y%m%d")}.pdf"'})


def _generate_excel_report(data):
    """Generate multi-sheet Excel report."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1E293B')
    sub_font = Font(name='Calibri', size=10, color='64748B')
    border = Border(bottom=Side(style='thin', color='E2E8F0'))
    crit_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    high_fill = PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid')
    med_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    pass_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fail_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    def style_header(ws, row=1, cols=10):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    ws.merge_cells('A1:F1')
    ws['A1'] = f"CloudSentinel Security Report — {data['account']}"
    ws['A1'].font = title_font
    ws['A2'] = f"Generated: {data['generated_at'][:19]}  |  Provider: {data['provider'].upper()}"
    ws['A2'].font = sub_font

    dash = data.get("dashboard", {})
    totals = dash.get("totals", {})
    row = 4
    ws.cell(row=row, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    style_header(ws, row, 2)
    metrics = [
        ("Security Score", f"{dash.get('security_score', 0)}/100"),
        ("WAF Overall Score", f"{data['waf'].get('overall_score', 0)}%"),
        ("Regions Scanned", dash.get('regions_scanned', 0)),
        ("EC2 Instances", totals.get('instances', 0)),
        ("S3 Buckets", totals.get('buckets', 0)),
        ("Security Groups", totals.get('security_groups', 0)),
        ("VPCs", totals.get('vpcs', 0)),
        ("Lambda Functions", totals.get('lambdas', 0)),
        ("RDS Instances", totals.get('rds', 0)),
        ("Snapshots", totals.get('snapshots', 0)),
        ("Subnets", totals.get('subnets', 0)),
        ("Network Interfaces", totals.get('network_interfaces', 0)),
        ("Public IPs", len(dash.get('public_ips', []))),
        ("Open Security Groups", dash.get('open_security_groups', 0)),
        ("IAM Users", dash.get('iam_summary', {}).get('Users', 0)),
        ("IAM Users without MFA", dash.get('iam_users_no_mfa', 0)),
        ("Total Audit Findings", data['audit']['total']),
        ("Critical Findings", data['audit']['summary'].get('CRITICAL', 0)),
        ("High Findings", data['audit']['summary'].get('HIGH', 0)),
        ("Medium Findings", data['audit']['summary'].get('MEDIUM', 0)),
    ]
    for i, (label, val) in enumerate(metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).border = border
        ws.cell(row=r, column=2, value=val).border = border
    auto_width(ws)

    # ── Sheet 2: Well-Architected Framework ──
    ws2 = wb.create_sheet("Well-Architected Framework")
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "AWS Well-Architected Framework Assessment"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Overall Score: {data['waf'].get('overall_score', 0)}%"
    ws2['A2'].font = Font(bold=True, size=12, color='4F46E5')

    row = 4
    headers = ["Pillar", "Score", "Passed", "Total", "Status"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, len(headers))

    for pid, pillar in data['waf'].get('pillars', {}).items():
        row += 1
        ws2.cell(row=row, column=1, value=pillar['name'])
        ws2.cell(row=row, column=2, value=f"{pillar['score']}%")
        ws2.cell(row=row, column=3, value=pillar['passed'])
        ws2.cell(row=row, column=4, value=pillar['total'])
        status = "PASS" if pillar['score'] >= 80 else "WARN" if pillar['score'] >= 50 else "FAIL"
        ws2.cell(row=row, column=5, value=status)
        ws2.cell(row=row, column=5).fill = pass_fill if status == "PASS" else med_fill if status == "WARN" else fail_fill
        for c in range(1, 6):
            ws2.cell(row=row, column=c).border = border

    row += 2
    ws2.cell(row=row, column=1, value="Detailed Checks").font = Font(bold=True, size=11)
    row += 1
    for c, h in enumerate(["Pillar", "Check", "Result", "Recommendation"], 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, 4)

    for pid, pillar in data['waf'].get('pillars', {}).items():
        for check in pillar.get('checks', []):
            row += 1
            ws2.cell(row=row, column=1, value=pillar['name'])
            ws2.cell(row=row, column=2, value=check['name'])
            ws2.cell(row=row, column=3, value="PASS" if check['passed'] else "FAIL")
            ws2.cell(row=row, column=3).fill = pass_fill if check['passed'] else fail_fill
            ws2.cell(row=row, column=4, value="" if check['passed'] else check['recommendation'])
            for c in range(1, 5):
                ws2.cell(row=row, column=c).border = border
    auto_width(ws2)

    # ── Sheet 3: Audit Findings ──
    ws3 = wb.create_sheet("Audit Findings")
    ws3['A1'] = f"Security Audit — {data['audit']['total']} Findings"
    ws3['A1'].font = title_font
    row = 3
    headers = ["Severity", "Title", "Issue ID", "Group", "Region", "Resource"]
    for c, h in enumerate(headers, 1):
        ws3.cell(row=row, column=c, value=h)
    style_header(ws3, row, len(headers))

    sev_fills = {"CRITICAL": crit_fill, "HIGH": high_fill, "MEDIUM": med_fill}
    for f in data['audit']['findings']:
        row += 1
        ws3.cell(row=row, column=1, value=f.get('severity', ''))
        ws3.cell(row=row, column=1).fill = sev_fills.get(f.get('severity', ''), PatternFill())
        ws3.cell(row=row, column=2, value=f.get('title', f.get('issue', '')))
        ws3.cell(row=row, column=3, value=f.get('issue', ''))
        ws3.cell(row=row, column=4, value=f.get('group', ''))
        ws3.cell(row=row, column=5, value=f.get('region', ''))
        ws3.cell(row=row, column=6, value=f.get('resource', ''))
        for c in range(1, 7):
            ws3.cell(row=row, column=c).border = border
    auto_width(ws3)

    # ── Sheet 4: IAM ──
    ws4 = wb.create_sheet("IAM Users & Roles")
    ws4['A1'] = "Identity & Access Management"
    ws4['A1'].font = title_font
    iam = data.get('iam', {})
    row = 3
    headers = ["Username", "ARN", "MFA Enabled", "Policies", "Groups"]
    for c, h in enumerate(headers, 1):
        ws4.cell(row=row, column=c, value=h)
    style_header(ws4, row, len(headers))
    for u in iam.get('users', []):
        row += 1
        ws4.cell(row=row, column=1, value=u.get('name', ''))
        ws4.cell(row=row, column=2, value=u.get('arn', ''))
        has_mfa = len(u.get('mfa_devices', [])) > 0
        ws4.cell(row=row, column=3, value="Yes" if has_mfa else "No")
        ws4.cell(row=row, column=3).fill = pass_fill if has_mfa else fail_fill
        ws4.cell(row=row, column=4, value=', '.join(u.get('policies', [])))
        ws4.cell(row=row, column=5, value=', '.join(u.get('groups', [])))
        for c in range(1, 6):
            ws4.cell(row=row, column=c).border = border
    auto_width(ws4)

    # ── Sheet 5: Security Groups ──
    ws5 = wb.create_sheet("Risky Security Groups")
    ws5['A1'] = "Security Groups Open to Internet"
    ws5['A1'].font = title_font
    row = 3
    headers = ["Severity", "Group Name", "Group ID", "Region", "VPC", "Protocol", "Port", "CIDR"]
    for c, h in enumerate(headers, 1):
        ws5.cell(row=row, column=c, value=h)
    style_header(ws5, row, len(headers))
    for sg in data.get('security_groups', {}).get('risky_groups', []):
        for rule in sg.get('open_rules', []):
            row += 1
            ws5.cell(row=row, column=1, value=sg.get('severity', ''))
            ws5.cell(row=row, column=1).fill = crit_fill if sg.get('severity') == 'CRITICAL' else high_fill
            ws5.cell(row=row, column=2, value=sg.get('group_name', ''))
            ws5.cell(row=row, column=3, value=sg.get('group_id', ''))
            ws5.cell(row=row, column=4, value=sg.get('region', ''))
            ws5.cell(row=row, column=5, value=sg.get('vpc_id', ''))
            proto = rule.get('protocol', '')
            ws5.cell(row=row, column=6, value='ALL' if proto == '-1' else proto.upper())
            ws5.cell(row=row, column=7, value=str(rule.get('from_port', 'ALL')))
            ws5.cell(row=row, column=8, value='0.0.0.0/0')
            for c in range(1, 9):
                ws5.cell(row=row, column=c).border = border
    auto_width(ws5)

    # ── Sheet 6: Region Matrix ──
    ws6 = wb.create_sheet("Region Matrix")
    ws6['A1'] = "Resource Distribution by Region"
    ws6['A1'].font = title_font
    rm = dash.get('region_matrix', {})
    if rm:
        res_keys = list(next(iter(rm.values())).keys()) if rm else []
        row = 3
        headers = ["Region"] + [k.replace('_', ' ').title() for k in res_keys] + ["Total"]
        for c, h in enumerate(headers, 1):
            ws6.cell(row=row, column=c, value=h)
        style_header(ws6, row, len(headers))
        for region, resources in sorted(rm.items(), key=lambda x: sum(x[1].values()), reverse=True):
            row += 1
            ws6.cell(row=row, column=1, value=region)
            total = 0
            for ci, key in enumerate(res_keys):
                val = resources.get(key, 0)
                total += val
                ws6.cell(row=row, column=ci + 2, value=val if val > 0 else "")
            ws6.cell(row=row, column=len(res_keys) + 2, value=total)
            for c in range(1, len(headers) + 1):
                ws6.cell(row=row, column=c).border = border
    auto_width(ws6)

    # ── Sheet 7: AI Recommendations ──
    ws7 = wb.create_sheet("AI Recommendations")
    ws7['A1'] = "CloudSentinel AI — Security Intelligence Report"
    ws7['A1'].font = title_font
    ws7['A2'] = f"WAF Score: {data['waf'].get('overall_score', 0)}%  |  Security Score: {dash.get('security_score', 0)}/100"
    ws7['A2'].font = Font(bold=True, size=11, color='4F46E5')
    row = 4
    for topic, content in data.get('ai_recommendations', {}).items():
        ws7.cell(row=row, column=1, value=content)
        ws7.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        ws7.row_dimensions[row].height = max(30, len(content.split('\n')) * 15)
        ws7.column_dimensions['A'].width = 120
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_comprehensive_pdf(data):
    """Generate comprehensive PDF report with all sections."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#1E293B'), spaceAfter=5)
    h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceBefore=15, spaceAfter=8)
    h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#334155'), spaceBefore=10, spaceAfter=5)
    body = ParagraphStyle('Body', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#475569'), leading=13)
    small = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94A3B8'))

    elements = []
    dash = data.get("dashboard", {})
    totals = dash.get("totals", {})

    # Title
    elements.append(Paragraph(f"CloudSentinel Security Report", title_style))
    elements.append(Paragraph(f"Account: {data['account']}  |  Provider: {data['provider'].upper()}  |  Generated: {data['generated_at'][:19]}", small))
    elements.append(Spacer(1, 10))

    # Executive Summary table
    elements.append(Paragraph("Executive Summary", h1))
    summary_data = [
        ["Security Score", f"{dash.get('security_score', 0)}/100", "WAF Score", f"{data['waf'].get('overall_score', 0)}%"],
        ["EC2 Instances", str(totals.get('instances', 0)), "S3 Buckets", str(totals.get('buckets', 0))],
        ["Security Groups", str(totals.get('security_groups', 0)), "Lambda Functions", str(totals.get('lambdas', 0))],
        ["VPCs", str(totals.get('vpcs', 0)), "RDS Instances", str(totals.get('rds', 0))],
        ["Open SGs", str(dash.get('open_security_groups', 0)), "Public IPs", str(len(dash.get('public_ips', [])))],
        ["IAM Users", str(dash.get('iam_summary', {}).get('Users', 0)), "Users w/o MFA", str(dash.get('iam_users_no_mfa', 0))],
        ["Total Findings", str(data['audit']['total']), "Critical", str(data['audit']['summary'].get('CRITICAL', 0))],
    ]
    t = Table(summary_data, colWidths=[80, 70, 80, 70])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748B')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#64748B')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('FONTNAME', (3, 0), (3, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    # WAF Section
    elements.append(Paragraph("Well-Architected Framework Assessment", h1))
    waf_rows = [["Pillar", "Score", "Passed", "Total", "Status"]]
    for pid, pillar in data['waf'].get('pillars', {}).items():
        status = "PASS" if pillar['score'] >= 80 else "WARN" if pillar['score'] >= 50 else "FAIL"
        waf_rows.append([pillar['name'], f"{pillar['score']}%", str(pillar['passed']), str(pillar['total']), status])
    waf_rows.append(["Overall", f"{data['waf'].get('overall_score', 0)}%", "", "", ""])

    t = Table(waf_rows, colWidths=[100, 50, 50, 50, 50])
    style_list = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
    ]
    for i, row in enumerate(waf_rows[1:-1], 1):
        if row[4] == "FAIL":
            style_list.append(('BACKGROUND', (4, i), (4, i), colors.HexColor('#FEE2E2')))
        elif row[4] == "WARN":
            style_list.append(('BACKGROUND', (4, i), (4, i), colors.HexColor('#FEF9C3')))
        elif row[4] == "PASS":
            style_list.append(('BACKGROUND', (4, i), (4, i), colors.HexColor('#D1FAE5')))
    t.setStyle(TableStyle(style_list))
    elements.append(t)

    # WAF Checks
    elements.append(Paragraph("Detailed WAF Checks", h2))
    for pid, pillar in data['waf'].get('pillars', {}).items():
        for check in pillar.get('checks', []):
            status = "✓" if check['passed'] else "✗"
            color = '#10B981' if check['passed'] else '#EF4444'
            text = f"<font color='{color}'>{status}</font> <b>{pillar['name']}</b> — {check['name']}"
            if not check['passed']:
                text += f" <font color='#94A3B8'>({check['recommendation']})</font>"
            elements.append(Paragraph(text, body))

    elements.append(PageBreak())

    # Audit Findings
    elements.append(Paragraph("Security Audit Findings", h1))
    elements.append(Paragraph(f"Total: {data['audit']['total']}  |  Critical: {data['audit']['summary'].get('CRITICAL',0)}  |  High: {data['audit']['summary'].get('HIGH',0)}  |  Medium: {data['audit']['summary'].get('MEDIUM',0)}", small))
    elements.append(Spacer(1, 5))

    audit_rows = [["Severity", "Title", "Region", "Resource"]]
    for f in data['audit']['findings'][:100]:
        audit_rows.append([
            f.get('severity', ''),
            (f.get('title', '') or f.get('issue', ''))[:60],
            f.get('region', '') or '',
            (f.get('resource', '') or '')[:40],
        ])
    t = Table(audit_rows, colWidths=[55, 180, 70, 120])
    style_list = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    sev_colors = {"CRITICAL": '#FEE2E2', "HIGH": '#FFEDD5', "MEDIUM": '#FEF9C3'}
    for i, row in enumerate(audit_rows[1:], 1):
        if row[0] in sev_colors:
            style_list.append(('BACKGROUND', (0, i), (0, i), colors.HexColor(sev_colors[row[0]])))
    t.setStyle(TableStyle(style_list))
    elements.append(t)

    elements.append(PageBreak())

    # AI Recommendations
    elements.append(Paragraph("AI Security Intelligence Report", h1))
    elements.append(Paragraph(f"CloudSentinel AI analysis based on your infrastructure scan data", small))
    elements.append(Spacer(1, 8))
    for topic, content in data.get('ai_recommendations', {}).items():
        for line in content.split('\n'):
            if line.startswith('## '):
                elements.append(Paragraph(line[3:], h2))
            elif line.startswith('**') and line.endswith('**'):
                elements.append(Paragraph(f"<b>{line.strip('*')}</b>", body))
            elif line.startswith('- '):
                elements.append(Paragraph(f"• {line[2:]}", body))
            elif line.strip():
                elements.append(Paragraph(line, body))

    doc.build(elements)
    return buf.getvalue()


# ── HEALTH ───────────────────────────────────────────────────────
@app.get("/api/health")
def health():
    accounts_with_data = _get_account_dirs()
    return {
        "status": "ok",
        "version": "3.0.0",
        "platform": "CloudSentinel Enterprise",
        "providers": registry.provider_ids,
        "accounts_with_data": accounts_with_data,
        "timestamp": datetime.now().isoformat(),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8088)
